from keras.models import load_model
from PIL import Image, ImageOps #Install pillow instead of PIL
import numpy as np
import openpyxl
import os
from PIL import Image

# Create an empty array to store the image characters
file_names = []

folder_path = "D:\Machine Learning\TMRC15_APAC_2\TMRC15_APAC_2"
folder_name = "TMRC15_APAC_2"
# Disable scientific notation for clarity
np.set_printoptions(suppress=True)

# Load the model
model = load_model('keras_Model.h5', compile=False)

# Load the labels
class_names= open('labels.txt', 'r').readlines()

model2 = load_model('keras_Model2.h5', compile=False)
# Load the labels
class_names2 = open('labels2.txt', 'r').readlines()


model_sensual = load_model('keras_model_sensual.h5', compile=False)

# Load the labels
class_names_sensual = open('labels_sensual.txt', 'r').readlines()

# Creat a class to store the image attributes
model_outdoor = load_model('keras_model_indoor.h5', compile=False)
class_names_outdoor = open('keras_model_indoor.txt', 'r').readlines()

# Creat a class to store the image attributes
model_face_noface = load_model('keras_model_face_noface.h5', compile=False)
class_names_face_noface = open('labels_face_noface.txt', 'r').readlines()

class ImageAttributes:
  def __init__(self, name, lowangle, phototype, sensual,outdoor,isface,):
    self.name = name
    self.isface = isface
    self.lowangle = lowangle
    self.phototype = phototype
    self.sensual =  sensual
    self.outdoor = outdoor

for images in os.listdir(folder_path):
 
    # check if the image ends with png
    if (images.endswith(".jpg") or images.endswith(".png")):
        print(images)
        folder_paths = "D:\Machine Learning\TMRC15_APAC_2\TMRC15_APAC_2"
        image = Image.open(folder_paths+"\\"+ images).convert('RGB')
    # Create the array of the right shape to feed into the keras model
        data = np.ndarray(shape=(1, 224, 224, 3), dtype=np.float32)
        data2 = np.ndarray(shape=(1, 224, 224, 3), dtype=np.float32)

    #resize the image to a 224x224 with the same strategy as in TM2:
    #resizing the image to be at least 224x224 and then cropping from the center
        size = (224, 224)
        image = ImageOps.fit(image, size, Image.Resampling.LANCZOS)

    #turn the image into a numpy array
        image_array = np.asarray(image)

    # Normalize the image
        normalized_image_array = (image_array.astype(np.float32) / 127.0) - 1
        normalized_image_array_2 = (image_array.astype(np.float32) / 127.0) - 1


    # Load the image into the array
        data[0] = normalized_image_array
        data2[0] = normalized_image_array_2


    # run the inference

        prediction_face = model_face_noface.predict(data)
        index_face = np.argmax(prediction_face)
        class_name_face = class_names_face_noface[index_face]
        confidence_score_face = prediction_face[0][index_face]

        prediction = model.predict(data)
        index = np.argmax(prediction)
        class_name = class_names[index]
        confidence_score = prediction[0][index]

        prediction2 = model2.predict(data)
        index2 = np.argmax(prediction2)
        class_name2 = class_names2[index2]
        confidence_score2 = prediction2[0][index2]

        prediction3 = model_sensual.predict(data)
        index3 = np.argmax(prediction3)
        class_name_sensual = class_names_sensual[index3]
        confidence_score_sensual = prediction3[0][index3]

        prediction4 = model_outdoor.predict(data2)
        index4 = np.argmax(prediction4)
        class_name_outdoor = class_names_outdoor[index4]
        confidence_score_outdoor = prediction4[0][index4]
        print('Class2:', class_name2, end='')
        print('Confidence score:', confidence_score2)

        print('Class:', class_name, end='')
        print('Confidence score:', confidence_score)

        print("Class Sensual:", class_name_sensual[2:], end="")
        print("Confidence Score:", confidence_score_sensual)

        print("Class Sensual:", class_name_outdoor[2:], end="")
        print("Confidence Score:",  confidence_score_outdoor)
        # file_names.append(ImageAttributes(images,class_name,class_name2,class_name_sensual[2:], class_name_outdoor[2:],class_name_face[:1]))

    # Add the image attributes to the image array based on the condition of face present
        if class_name_face[:1] =='0':
          file_names.append(ImageAttributes(images,"-","-","-", "-", class_name_face))
          print("This doesnt has any face", class_name_face)

        if class_name_face[:1] =='1':
          print("This has face", class_name_face)
          file_names.append(ImageAttributes(images,class_name,class_name2,class_name_sensual[2:], class_name_outdoor[2:],class_name_face[:1]))


# Print the array of file names
print(file_names)

workbook = openpyxl.Workbook()

# Get the active sheet
sheet = workbook.active
sheet.cell(row=1, column= 1, value= "Image Name")
sheet.cell(row=1, column= 3, value= "Folder Name")
sheet.cell(row=1, column= 5, value= "Sensuality")
sheet.cell(row=1, column= 6, value= "Indoor/Outdoor")


for i, obj in enumerate(file_names):
    print(obj.name)
    sheet.cell(row=i+2, column= 1, value=obj.name)
    sheet.cell(row=i+2, column= 3, value=folder_name)
    sheet.cell(row=i+2, column= 5, value=obj.sensual)
    sheet.cell(row=i+2, column= 6, value=obj.outdoor)
    sheet.cell(row=i+2, column= 15, value=obj.isface)


# Save the workbook to an Excel file
workbook.save(  "codesbook.xlsx")

